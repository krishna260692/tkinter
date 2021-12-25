from tkinter import *
import sqlite3
import tkinter.messagebox as messagebox
from tkinter import ttk
from time import strftime
import openpyxl
from openpyxl import load_workbook
import os

import sys
sys.setrecursionlimit(1000000)

if os.environ.get('DISPLAY','') == '':
    print('no display found. Using :0.0')
    os.environ.__setitem__('DISPLAY', ':0.0')


root=Tk()
root.geometry("1350x700")
#root.resizable(height = None, width = None)
root.title("Payroll Solution")
root.iconbitmap("C:\\Users\\krish\\Downloads\\ico.ico")

NAME = StringVar()
EMPLOYEE = StringVar()
DESIGNATION = StringVar()
DOJ = StringVar()
DEPARTMENT = StringVar()
BASICSALARY =IntVar()
TOTALWORKING= IntVar()

########################################################## function part start here #####################################################################



def home():
    # frame.destroy()
    topl.destroy()
def home1():
    # frame.destroy()
    # frame1.destroy()
    topl2.destroy()
def home2():
    root.destroy()



def sqlentry():

    conn = sqlite3.connect("sqldata.db")
    cursor = conn.cursor()
    #cursor.execute('DROP TABLE employee')
    cursor.execute("CREATE TABLE IF NOT EXISTS employee(id INTEGER PRIMARY KEY AUTOINCREMENT NOT NULL, name TEXT, employee TEXT, designation TEXT, doj TEXT, department TEXT ,Basic_Salary TEXT, Total_working TEXT, Earning_Salary TEXT)")
    conn.commit()
    conn.close()
sqlentry()

def add():
    global d
    global z
    a = BASICSALARY.get()
    b = TOTALWORKING.get()
    c = a / 30
    d = c * b


    conn = sqlite3.connect("sqldata.db")
    cursor = conn.cursor()
    cursor.execute("INSERT INTO employee (name,  employee, designation, doj, department, Basic_Salary, Total_working, Earning_Salary) VALUES (?,?, ?,?,?,?,?,?)", (str(NAME.get()),
                      str(EMPLOYEE.get()), str(DESIGNATION.get()), str(DOJ.get()),str(DEPARTMENT.get()),int(BASICSALARY.get()), int(TOTALWORKING.get()) ,d))
    conn.commit()
    print("INSERTED")
    conn.commit()

    wb=load_workbook("Book1.xlsx")
    ws=wb.active
    ws1 = NAME.get()
    ws2 = EMPLOYEE.get()
    ws3= DESIGNATION.get()
    ws4 = DOJ.get()
    ws5 = DEPARTMENT.get()
    ws6= BASICSALARY.get()
    ws7= TOTALWORKING.get()
    ws8= d
    ws.append([ws1,ws2,ws3,ws4,ws5,ws6,ws7,ws8])

    wb.save("Book1.xlsx")
    messagebox._show('Message', 'Saved Successfully')
    topl.destroy()




def fetch():
    conn = sqlite3.connect("sqldata.db")
    cur = conn.cursor()
    cur.execute("SELECT * FROM employee")
    row = cur.fetchall()
    if len(row) != 0:
        table.delete(*table.get_children())
        for rows in row:
            table.insert('', END, values=rows)
        conn.commit()
def delete():
    a = DELETEID.get()
    conn = sqlite3.connect("sqldata.db")
    cur = conn.cursor()
    cur.execute("DELETE FROM employee WHERE id =?", (a,))
    conn.commit()
    conn.close()
    entry.delete(0, END)
    messagebox._show('Message', 'Deleted')

def cleartext():
    entry1.delete(0, END)
    entry2.delete(0, END)
    entry3.delete(0, END)
    entry4.delete(0, END)
    entry5.delete(0, END)
    entry6.delete(0, END)
    entry7.delete(0, END)

def droptable():
    result = messagebox.askquestion('Management System', 'DO YOU WANT TO DROP TABLE', icon='warning')
    if result == 'yes':

        conn = sqlite3.connect("sqldata.db")
        cursor = conn.cursor()
        cursor.execute('DROP TABLE employee')
        conn.commit()
        conn.close()
        messagebox._show('Message', 'Table Deleted')


def salary():
    name= "Book1.xlsx"
    os.system(name)

def check():
    book = load_workbook('Book1.xlsx')
    sheet = book.active

    print("Maximum rows before removing:", sheet.max_row)
    sheet.delete_rows(2, sheet.max_row)
    print("Maximum rows after removing:", sheet.max_row)


    book.save('Book1.xlsx')

check()



def fetchdb():
    topl2 = Toplevel()
    topl2.geometry("1200x600")
    topl2.resizable(width=True, height=True)



    conn= sqlite3.connect("sqldata.db")
    c=conn.cursor()
    c.execute("SELECT * FROM employee")
    r=c.fetchall()

    for i in r:

        num =2
        name= Label(topl2,  text = i[0], font = "time 12 bold", fg= "blue")
        name.place(x=num, y=20)

        empc = Label(topl2, text=i[1], font="time 12 bold", fg="blue")
        empc.place(x=num, y=20)

        dest = Label(topl2, text=i[2], font="time 12 bold", fg="blue")
        dest.place(x=200, y=20)

        doj = Label(topl2, text=i[3], font="time 12 bold", fg="blue")
        doj.place(x=250, y=20)

        dep = Label(topl2, text=i[4], font="time 12 bold", fg="blue")
        dep.place(x=350, y=20)

        basics = Label(topl2, text=i[5], font="time 12 bold", fg="blue")
        basics.place(x=450, y=20)

        totalw = Label(topl2, text=i[6], font="time 12 bold", fg="blue")
        totalw.place(x=500, y=20)

        earning = Label(topl2, text=i[7], font="time 12 bold", fg="blue")
        earning.place(x=650, y=20)

        num = num+1




















    topl2.mainloop()




########################################################################### ADD EMPLOYEE FUNCTION  ###########################################################

def addemployee():


    global frame ,topl
    topl=Toplevel()
    topl.geometry("600x500")
    topl.resizable(width=False, height=False)
    topl.config(bg="light blue")
    # frame= Frame(topl,bd=2, bg='tan3', relief=RIDGE)
    # frame.place(x=0,y=0,width=600,height=500)

    button = Button(topl, text='HOME', width=20, height=2, bg='light coral',activebackground='purple1', command=home)
    button.place(x=20, y=20)
    button = Button(topl, text='ADD', width=20, height=2, bg='light coral',activebackground='purple1', command=add)
    button.place(x=200, y=20)
    button = Button(topl, text='CLEAR', width=20, height=2, bg='light coral', activebackground='purple1', command=cleartext)
    button.place(x=380, y=20)

    label = Label(topl, text='Name', height=1, font=('Lucida Fax', 15), bg='light blue')
    label.place(x=10, y=100)
    label = Label(topl, text='Employee Code', height=1, font=('Lucida Fax', 15), bg='light blue')
    label.place(x=10, y=140)
    label = Label(topl, text='Designation', height=1, font=('Lucida Fax', 15), bg='light blue')
    label.place(x=10, y=180)
    label = Label(topl, text='D.O.J', height=1, font=('Lucida Fax', 15), bg='light blue')
    label.place(x=10, y=220)
    label = Label(topl, text='Department', height=1, font=('Lucida Fax', 15), bg='light blue')
    label.place(x=10, y=260)
    label = Label(topl, text='Basic Salary', height=1, font=('Lucida Fax', 15), bg='light blue')
    label.place(x=10, y=300)
    label = Label(topl, text='Total Working\n Days', height=2, font=('Lucida Fax', 15), bg='light blue')
    label.place(x=10, y=340)

    global entry1
    global entry2
    global entry3
    global entry4
    global entry5
    global entry6
    global entry7

    entry1 = Entry(topl, textvariable=NAME, font=('Lucida Fax', 15))
    entry1.place(x=200, y=100)
    entry2 = Entry(topl, textvariable=EMPLOYEE, font=('Lucida Fax', 15))
    entry2.place(x=200, y=140)
    entry3 = Entry(topl, textvariable=DESIGNATION, font=('Lucida Fax', 15))
    entry3.place(x=200, y=180)
    entry4 = Entry(topl, textvariable=DOJ, font=('Lucida Fax', 15))
    entry4.place(x=200, y=220)
    entry5 = Entry(topl, textvariable=DEPARTMENT, font=('Lucida Fax', 15))
    entry5.place(x=200, y=260)
    entry6 = Entry(topl, textvariable=BASICSALARY, font=('Lucida Fax', 15))
    entry6.place(x=200, y=300)
    entry7 = Entry(topl, textvariable=TOTALWORKING, font=('Lucida Fax', 15))
    entry7.place(x=200, y=340)

    topl.mainloop()

DELETEID= StringVar()

############################################################ SHOW DATA BUTTON FUNCTION ############################################################
def show():

    global show
    global frame
    global frame1
    global table
    global DELETEID
    global entry
    global topl2
    topl2 =Toplevel()
    topl2.geometry("1200x600")
    topl2.resizable(width=False, height=False)
    frame1 = Frame(topl2, bd=2,bg='white' ,relief=RIDGE)
    frame1.place(x=0, y=0, width=295, height=700)
    frame= Frame(topl2,bd=2, relief=RIDGE)
    frame.place(x=300,y=0,width=900,height=680)

    button = Button(frame1, text='HOME', width=20, height=2, bg='coral', command=home1)
    button.place(x=10, y=20)
    button = Button(frame1, text='SHOW', width=20, height=2, bg='coral', command=fetch)
    button.place(x=10, y=70)
    button = Button(frame1, text='DELETE', width=20, height=2, bg='coral', command=delete)
    button.place(x=10, y=120)
    label =Label(frame1,text ='Enter Id To Delete',bg='white',font=('caliberi',12))
    label.place(x=10,y=180)
    entry=Entry(frame1,font=('caliberi',12),textvariable=DELETEID,bg='light gray')
    entry.place(x=10,y=210,width=120,height=25)

    scroll_x = ttk.Scrollbar(frame, orient=HORIZONTAL)
    scroll_y = ttk.Scrollbar(frame, orient=VERTICAL)

    table = ttk.Treeview(frame, columns=("id", "name",  "employee", "designation", "doj", "department", "Basic_Salary", "Total_working", "Earning_Salary" ), xscrollcommand=scroll_x.set, yscrollcommand=scroll_y.set)
    scroll_x.pack(side=BOTTOM, fill=X)
    scroll_y.pack(side=RIGHT, fill=Y)
    scroll_x.config(command=table.xview)
    scroll_y.config(command=table.yview)
    table.heading("id", text='ID')
    table.heading("name", text='NAME')
    table.heading("employee", text='EMPLOYEE CODE')
    table.heading("designation", text='DESIGNATION')
    table.heading('doj', text='D.O.J')
    table.heading('department', text='DEPARTMENT')
    table.heading('Basic_Salary', text='SALARY')
    table.heading('Total_working', text='TOTAL WORKING DAYS')
    table.heading('Earning_Salary', text='TOTAL SALARY')
    table['show'] = 'headings'

    table['show'] = 'headings'
    table.column('id', width=30)
    table.column("name", width=100)
    table.column("employee", width=100)
    table.column("designation", width=120)
    table.column("doj", width=100)
    table.column("department", width=100)
    table.column("Basic_Salary", width=120)
    table.column("Total_working", width=100)
    table.column("Earning_Salary", width=100)


    table.pack(fill=BOTH, expand=1)
    #table.bind("<ButtonRelease-1>", getcurser)
    topl2.mainloop()

################################################################## CLOCK FUNCTION   ###################################################
frame2= Frame(root, bg='white')
frame2.place(x=1200,y=0,width=150, height=30)

#frame2.title("clock")
def time():
    string= strftime("%H:%M:%S %p")
    label.config(text=string)
    label.after(1000,time)

label = Label (frame2, font=("ds-digital",18),background="white",foreground='black')
label.pack(anchor='center')
time()

import datetime

x = datetime.datetime.now()
print(x)

################################################################### button part start here ###################################################

button= Button(root, text= 'ADD EMPLOYEE', width=20, height=2,fg='black', bg='cyan3', font=('Bodoni MT',10), command=addemployee)
button.place(x=20, y=70)
button= Button(root, text= 'DROP TABLE', width=20, height=2,fg='black', bg='cyan3', font=('Bodoni MT',10), command=droptable)
button.place(x=210, y=70)
button= Button(root, text= 'DROP EXECL_FILE', width=20, height=2,fg='black', bg='coral', font=('Bodoni MT',10), command=check)
button.place(x=210, y=120)

button= Button(root, text= 'CLOSE', width=20, height=2, bg='red2',fg='white',  font=('Bodoni MT',10), command=home2)
button.place(x=20, y=270)

button= Button(root, text= 'SHOW DATA', width=20, height=2, bg='coral', font=('Bodoni MT',10), command=show)
button.place(x=20, y=170)

button= Button(root, text= 'EXPORT TO EXCEL', width=20, height=2, bg='coral', font=('Bodoni MT',10) ,command=salary)
button.place(x=20, y=120)

button= Button(root, text= 'PENDING ', width=20, height=2, bg='coral', font=('Bodoni MT',10) ,command='')
button.place(x=20, y=220)

button= Button(root, text= ' PENDING', width=20, height=2, bg='coral', font=('Bodoni MT',10) ,command='')
button.place(x=210, y=170)



root.mainloop()


